import csv
import CONST_COLUMN as CONST
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path
from typing import List, Dict

from SSLRequest import get_certificate_data
from UI import change_progress_ui
from CONST_COLUMN import WARNING_TIME_CONST, SUCCESS_TIME_CONST

def load_csv(file):
    original_path = Path(file)

    if not original_path.is_file():
        raise FileNotFoundError(f"Файл не найден: {file}")

    #return csv_to_excel_manual(original_path)
    return original_path

def get_csv_data_dict(copy_file_path):
    with open(copy_file_path, mode='r', encoding='cp1251', newline='') as f:
        reader = csv.DictReader(f,delimiter=';')
        data = list(reader)  # список словарей
        return data

def get_new_csv_data_with_ssl(list_of_dicts):
    i = 0
    for dictionary in list_of_dicts:

        if dictionary[CONST.NAME] is None:
            break

        ssl_data = get_certificate_data(dictionary[CONST.NAME])
        if ssl_data is not None:
            dictionary[CONST.DOMAIN_NAME] = ssl_data.domain
            dictionary[CONST.END_DATE] = ssl_data.end_date
            print(ssl_data.domain)
        else:
            dictionary[CONST.INFO] = CONST.SSL_ERROR_TEMPLATE
        i+=1
        change_progress_ui(len(list_of_dicts), i)

    return list_of_dicts

# def save_csv(list_of_dicts, file_path):
#     # Берём заголовки из ключей первого словаря
#     fieldnames = list_of_dicts[0].keys()
#
#     with open(file_path, mode='w', encoding='utf-8-sig', newline='') as f:
#         writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
#         writer.writeheader()
#         writer.writerows(list_of_dicts)

# def csv_to_excel_manual(csv_file: str, excel_file: str = None):
#     if excel_file is None:
#         excel_file = Path(csv_file).with_suffix('.xlsx')
#
#     wb = Workbook()
#     ws = wb.active
#
#     with open(csv_file, 'r', encoding='utf-8-sig', newline='') as f:
#         reader = csv.reader(f, delimiter=';')
#         for row in reader:
#             ws.append(row)
#
#     wb.save(excel_file)
#     return excel_file

def save_colored_excel(list_of_dicts: List[Dict], output_path: Path):
    if not list_of_dicts:
        raise ValueError("Нет данных для сохранения")

    output_file_path = output_path.parent / "output.xlsx"

    # Создаём новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты SSL"

    # Заголовки
    headers = list(list_of_dicts[0].keys())
    ws.append(headers)

    # Стили
    error_fill = PatternFill(start_color="FFBDAD", end_color="FFBDAD", fill_type="solid")  # Красный фон (ошибка)
    success_fill = PatternFill(start_color="ABF5D1", end_color="ABF5D1", fill_type="solid")  # Зелёный фон (успех)
    warning_fill = PatternFill(start_color="FFF0B3", end_color="FFF0B3", fill_type="solid")
    bold_font = Font(bold=True)

    # Форматируем заголовки
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = bold_font

    # Заполняем строки и применяем цвета
    for row_idx, row_data in enumerate(list_of_dicts, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Пример: раскраска всей строки, если в колонке INFO есть ошибка
            if header == CONST.INFO and CONST.SSL_ERROR_TEMPLATE in str(value):
                # Раскрасить ВСЮ строку в красный
                # for c in range(1, len(headers) + 1):
                #     ws.cell(row=row_idx, column=c).fill = error_fill
                # break
                cell.fill = error_fill
            elif header == CONST.END_DATE and value:
                # Можно раскрасить только ячейку с датой, например, зелёным
                code = checkValidDateTime(value)
                if code == 0:
                    cell.fill = success_fill
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=c).fill = success_fill
                elif code == 1:
                    cell.fill = success_fill
                elif code == 2:
                    cell.fill = warning_fill
                else:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=c).fill = error_fill

    # Автоподбор ширины колонок (опционально)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем
    wb.save(output_file_path)
    print(f"Результат сохранён в: {output_file_path}")
    return output_file_path

    # 0 - success > suc_const, 1 - success < suc_const, 2 - warning, 3 - expired
def checkValidDateTime(date_str):
    ssl_date = dt.datetime.strptime(date_str, "%d.%m.%Y")
    today = dt.datetime.today()
    delta = ssl_date - today
    print(delta)
    if delta.days > SUCCESS_TIME_CONST:
        return 0
    elif SUCCESS_TIME_CONST >= delta.days > WARNING_TIME_CONST:
        return 1
    elif 0 <= delta.days <= WARNING_TIME_CONST:
        return 2
    else:
        return 3
